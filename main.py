import json
import os
import pprint
import time
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Color, fonts
import matplotlib.pyplot as plt
from decimal import getcontext, Decimal
from urllib.request import urlopen

getcontext().prec = 6


def write_workbook(save_location: str):  # testing

    # 这是一个应急功能 着急的时候没人care图和excel
    def rankings_excel(teams: list, season: str, start_row: int = 1, start_column: int = 1):
        rankings_columns = ("Team", "Wins", "Losses", "AP", "Ranking", "Highest", "Result")
        for x, y in enumerate(rankings_columns):  # Initialize Matches
            book[sheet_names[1]].cell(row=1, column=x + 1).value = y
            book[sheet_names[1]].cell(row=1, column=x + 1).font = ExcelStyle.BOLD_BLACK_FONT

        for row, team in enumerate(rankings_scan(teams, season)):
            for column, value in enumerate(team):
                book[sheet_names[1]].cell(row=row + start_row, column=6 + start_column).value = value
            if int(team[1]) > int(team[2]):
                book[sheet_names[1]].cell(row=row + start_row + 1, column=6 + start_column).value = "Positive"
                book[sheet_names[1]].cell(row=row + start_row + 1, column=6 + start_column).fill = ExcelStyle.RED_FILL
            elif int(team[1]) < int(team[2]):
                book[sheet_names[1]].cell(row=row + start_row + 1, column=6 + start_column).value = "Negative"
                book[sheet_names[1]].cell(row=row + start_row + 1, column=6 + start_column).fill = ExcelStyle.BLUE_FILL
            elif int(team[1]) == int(team[2]):
                book[sheet_names[1]].cell(row=row + start_row + 1, column=6 + start_column).value = "Equal"
                book[sheet_names[1]].cell(row=row + start_row, column=6 + start_column).fill = ExcelStyle.BLACK_FILL
            else:
                book[sheet_names[1]].cell(row=row + start_row + 1, column=6 + start_column).value = "Error"
                book[sheet_names[1]].cell(row=row + start_row + 1, column=6 + start_column).fill = ExcelStyle.GREEN_FILL
            book[sheet_names[1]].cell(row=row + 1, column=6 + start_column).font = ExcelStyle.BOLD_WHITE_FONT

    def matches_excel(team: str, season: str, start_row: int = 1, start_column: int = 1):
        matches_columns = (
            "Sku", "Match", "Red1", "Red2", "Red3", "RedSit", "Blue1", "Blue2", "Blue3", "BlueSit", "RedSco", "BlueSco"
        )
        for x, y in enumerate(matches_columns):  # Initialize Matches
            book[sheet_names[1]].cell(row=1, column=x + 1).value = y
            book[sheet_names[1]].cell(row=1, column=x + 1).font = ExcelStyle.BOLD_BLACK_FONT

        for row, match in enumerate(matches_scan(team, season)):
            for column, value in enumerate(match):
                book[sheet_names[1]].cell(row=start_row + row, column=start_column + column)
            if int(match[10]) > int(match[11]):
                book[sheet_names[1]].cell(row=row + 1, column=start_column + 14).value = "Red"
                book[sheet_names[1]].cell(row=row + 1, column=start_column + 14).fill = ExcelStyle.RED_FILL
            elif int(match[10]) < int(match[11]):
                book[sheet_names[1]].cell(row=row + 1, column=start_column + 14).value = "Blue"
                book[sheet_names[1]].cell(row=row + 1, column=start_column + 14).fill = ExcelStyle.BLUE_FILL
            if int(match[10]) + 20 < int(match[11]):
                book[sheet_names[1]].cell(row=row + 1, column=start_column + 14).value = "Blue Easy"
                book[sheet_names[1]].cell(row=row + 1, column=start_column + 14).fill = ExcelStyle.GREEN_FILL
            elif int(match[10]) - 20 > int(match[11]):
                book[sheet_names[1]].cell(row=row + 1, column=start_column + 14).value = "Red Easy"
                book[sheet_names[1]].cell(row=row + 1, column=start_column + 14).fill = ExcelStyle.YELLOW_FILL

            book[sheet_names[1]].cell(row=row + 1, column=start_column + 14).font = ExcelStyle.BOLD_WHITE_FONT

            if match[2] == team or match[3] == team or match[4] == team:
                if int(match[10]) > int(match[11]):
                    book[sheet_names[1]].cell(row=row + 1, column=start_column + 13).value = "Win"
                    book[sheet_names[1]].cell(row=row + 1, column=start_column + 13).font = ExcelStyle.BOLD_BLACK_FONT
                else:
                    book[sheet_names[1]].cell(row=row + 1, column=start_column + 13).value = "Lose"
                    book[sheet_names[1]].cell(row=row + 1, column=start_column + 13).font = ExcelStyle.BOLD_WHITE_FONT
                    book[sheet_names[1]].cell(row=row + 1, column=start_column + 13).fill = ExcelStyle.BLACK_FILL
            elif match[6] == team or match[7] == team or match[8] == team:
                if int(match[10]) < int(match[11]):
                    book[sheet_names[1]].cell(row=row + 1, column=start_column + 13).value = "Win"
                    book[sheet_names[1]].cell(row=row + 1, column=start_column + 13).font = ExcelStyle.BOLD_BLACK_FONT
                else:
                    book[sheet_names[1]].cell(row=row + 1, column=start_column + 13).value = "Lose"
                    book[sheet_names[1]].cell(row=row + 1, column=start_column + 13).font = ExcelStyle.BOLD_WHITE_FONT
                    book[sheet_names[1]].cell(row=row + 1, column=start_column + 13).fill = ExcelStyle.BLACK_FILL

    class ExcelStyle:
        RED_FILL = PatternFill(patternType="solid", fgColor=colors.RED)
        BLUE_FILL = PatternFill(patternType="solid", fgColor=colors.BLUE)
        GREEN_FILL = PatternFill(patternType="solid", fgColor=colors.GREEN)  # replace Light Red
        YELLOW_FILL = PatternFill(patternType="solid", fgColor=colors.YELLOW)  # replace Light_Blue
        BLACK_FILL = PatternFill(patternType="solid", fgColor=colors.BLACK)
        BOLD_RED_FONT = Font("Calibri", size=11, color=colors.RED, bold=True)
        BOLD_BLUE_FONT = Font("Calibri", size=11, color=colors.BLUE, bold=True)
        BOLD_BLACK_FONT = Font("Calibri", size=11, color=colors.BLACK, bold=True)
        BOLD_WHITE_FONT = Font("Calibri", size=11, color=colors.WHITE, bold=True)

    # Initialize the workbook
    book = openpyxl.Workbook()
    sheet_names = ("#Cover", "#Rankings", "#Important Data", "#For World", "#Bugged Teams")

    for x in sheet_names:
        book.create_sheet(x)
    del book["Sheet"]  # I don't know how to solve this myth, it automatically generate sheets
    book[sheet_names[0]].cell(row=1, column=1).value = "Last Change:" + str(time.localtime())
    rankings_excel(teams=["35211C"], season="Turning%20Point")
    book.save(save_location)
    return True

#             "Because of there are no data for these teams: 1119S, 7386A, 8000X, 8000Z, 19771B, 30638A, 36632A, "
#            "37073A, 60900A, 76921B, 99556A, 99691E, 99691H are not include in the sheet #Important Data")

def vexdb_json(api_type: str, api_parameters: dict, return_data=None):
    """
    It function accept a string "api_type" and a dictionary "api_parameters", the "api_type" should be
    one from _API_TYPE The dictionary's key are the _parameters from vexdb.io/the_data and the value should
    also follow it.
    """
    # TODO(Yifei): Multi thread, timeout retry,throw error correctly

    if return_data is None:
        return_data = ["full"]
    _parameters = ""
    output = None

    if api_parameters:
        if type(api_parameters) == dict:
            _keys = list(api_parameters.keys())
            _values = list(api_parameters.values())
            if len(_keys) >= 1:
                _parameters += "?" + _keys[0] + "=" + _values[0]
                if len(_keys) > 1:
                    for x in range(1, len(_keys)):
                        _parameters += "&" + _keys[x] + "=" + _values[x]
    else:
        _parameters = None

    if api_type != "":
        if _parameters != "" or _parameters is not None:
            json_dict = json.loads((urlopen("http://api.vexdb.io/v1/get_" + api_type + _parameters)).read())
            if json_dict["status"] == 0:
                raise (IOError)  # TODO: a exception
            else:
                if json_dict["size"] == 5000:
                    raise (IOError)  # TODO: Another exception or use some trick to prevent 5000 limit
                else:
                    if return_data[0] == "full":
                        output: dict = json_dict
                    if return_data[0] != "full":
                        output: list = []
                        for x in json_dict["result"]:
                            for y in return_data:
                                output.append(x[y])
                return output


def team_list():  # For testing
    #print(vexdb_json("teams", {"grade": "High%20School"}, ["number"]))
    print(vexdb_json("matches", {"season": "Starstruck", "team": "8667A"}, ["sku"]))


def matches_scan(team: str, season: str):
    out = []
    for r in vexdb_json("matches", {"season": season, "team": team}):
        out.append([str(r["sku"]), str(r["matchnum"]), str(r["round"]),
                    str(r["red1"]), str(r["red2"]), str(r["red3"]),
                    str(r["redsit"]), str(r["blue1"]), str(r["blue2"]),
                    str(r["blue3"]), str(r["bluesit"]), str(r["redscore"]),
                    str(r["bluescore"])])
    return out


def rankings_scan(teams: list, season: str):
    out = []
    for x, team in enumerate(teams):
        for r in vexdb_json("rankings", {"team": team, "season": season})["result"]:
            a = {"team": r["team"], "wins": r["wins"], "losses": r["losses"], "ap": r["ap"], "rank": r["rank"], "max_score":r["max_score"]}
            out.append(a)
    return out


def getteam(sku, country):
    # TODO: fix after finish readconfig
    _json_dict = vexdb_json("teams", {"sku": sku, "program": "VRC", "limit_number": "4999", "country": country})
    output = []
    for r in _json_dict["result"]:
        line = '{}: '.format(r["number"])
        output.append(line)
    return output


def main():
    write_workbook("./test.xlsx")


if __name__ == '__main__':
    main()
