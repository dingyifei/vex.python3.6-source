"""
This is a rewrite of the vexdb excel visualizer module, it don't need internet connection if you use offline data
Credit:Haorui Zhou (oraginal code), Yifei Ding(Rewrite)
license: CC By-NC-SA
contact: yifeiding@protonmail.com
"""
import time

import openpyxl
from openpyxl.styles import PatternFill, Font, colors


class WriteWorkbook:
    """
    if you see this, you better check each function's docstring because I don't know how to explain this
    """

    def __init__(self):

        self.save_location = "./output.xlsx"
        self.RED_FILL = PatternFill(patternType="solid", fgColor=colors.RED)
        self.BLUE_FILL = PatternFill(patternType="solid", fgColor=colors.BLUE)
        self.GREEN_FILL = PatternFill(patternType="solid", fgColor=colors.GREEN)  # replace Light Red
        self.YELLOW_FILL = PatternFill(patternType="solid", fgColor=colors.YELLOW)  # replace Light_Blue
        self.BLACK_FILL = PatternFill(patternType="solid", fgColor=colors.BLACK)
        self.WHITE_FILL = PatternFill(patternType="solid", fgColor=colors.WHITE)
        self.BOLD_RED_FONT = Font("Calibri", size=11, color=colors.RED, bold=True)
        self.BOLD_BLUE_FONT = Font("Calibri", size=11, color=colors.BLUE, bold=True)
        self.BOLD_BLACK_FONT = Font("Calibri", size=11, color=colors.BLACK, bold=True)
        self.BOLD_WHITE_FONT = Font("Calibri", size=11, color=colors.WHITE, bold=True)
        self.book = openpyxl.Workbook()
        self.book.create_sheet("Cover")
        del self.book["Sheet"]  # I don't know how to solve this myth, it automatically generate sheets
        self.book["Cover"].cell(row=1, column=1).value = "Last Change:" + str(time.localtime())

    @staticmethod
    def value_check(values: list):
        """
        it check the 2d list is the correct format
        :param values: the 2d list, it will make sure the 2d list is correct format
        """
        if type(values) is list:
            if len(values) == 3:
                if type(values[0]) != str:
                    raise ValueError("The first value in list should be string")
                if type(values[1]) != PatternFill:
                    raise ValueError("The second value in list should be PatternFill")
                if type(values[2]) != Font:
                    raise ValueError("The third value in list should be Font")
            else:
                raise ValueError("invalid tuple length")
        else:
            raise ValueError("The value should be tuple contain two value")

    def write_chart(self, sheet: str, text: list, start_row=1, start_column=1):
        """
        This function write chart, YOU MUST USE THE CORRECT FORMAT 2D LIST!!!!
        :param sheet: if the sheet does not exit, it will be created automatically
        :param text: a 2d list, for example: [[{"test": (a.YELLOW_FILL, a.BOLD_BLUE_FONT)}]]
        :param start_row: the start row, should be the top row
        :param start_column: the start column, should be the left column
        """
        if sheet not in self.book.sheetnames:
            self.book.create_sheet(sheet)
        active_sheet = self.book[sheet]
        for row, a in enumerate(text):
            for column, b in enumerate(a):
                self.value_check(b)
                active_sheet.cell(row=start_row + row, column=start_column + column).value = b[0]
                active_sheet.cell(row=start_row + row, column=start_column + column).fill = b[1]
                active_sheet.cell(row=start_row + row, column=start_column + column).font = b[2]

    def list_formater(self, text: list,
                      font=Font("Calibri", size=11, color=colors.BLACK, bold=True),
                      fill=PatternFill(patternType="solid", fgColor=colors.BLACK)
                      ):
        out = []
        for row, a in enumerate(text):
            add = []
            for column, b in enumerate(a):
                add.append([b, fill, font])
            out.append(add)
        return out

    def save(self):
        """
        it save the file
        """
        self.book.save(self.save_location)


#  there are no data for these teams:
#  1119S, 7386A, 8000X, 8000Z, 19771B, 30638A, 36632A 37073A, 60900A, 76921B, 99556A, 99691E, 99691H for some reason


def main():
    """
    usually this main doesn't do anything
    """
    a = WriteWorkbook()
    print(a.list_formater([["testing"]]))
    print(a.BOLD_BLACK_FONT)
    a.write_chart("test", [[["testing", a.YELLOW_FILL, a.BOLD_BLUE_FONT]]])
    a.save()


if __name__ == '__main__':
    main()
